"""Reusable openpyxl styling helpers."""

from __future__ import annotations

from collections.abc import Mapping

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from mission_control.core.theme import Theme


def solid_fill(color: str) -> PatternFill:
    """Create a solid fill from a hex color without a leading hash."""
    return PatternFill(fill_type="solid", fgColor=color)


def thin_border(color: str) -> Border:
    """Return a thin border for table cells."""
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def apply_table_header_style(
    worksheet: Worksheet,
    row: int,
    start_column: int,
    end_column: int,
    theme: Theme,
) -> None:
    """Apply the standard table header style to a contiguous row range."""
    for column in range(start_column, end_column + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.fill = solid_fill(theme.primary)
        cell.font = Font(
            name=theme.header_font,
            size=theme.header_size,
            bold=True,
            color=theme.white,
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border(theme.border)


def apply_column_widths(
    worksheet: Worksheet,
    widths: Mapping[str, float],
) -> None:
    """Apply explicit widths to worksheet columns."""
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def auto_size_columns(
    worksheet: Worksheet,
    *,
    min_width: float = 12,
    max_width: float = 34,
    padding: float = 2,
) -> None:
    """Auto-size populated columns within sensible bounds."""
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))

        if max_length:
            width = min(max(max_length + padding, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = width
