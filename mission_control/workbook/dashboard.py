"""Dashboard worksheet implementation."""

from __future__ import annotations

from datetime import date
from typing import Any

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font

from mission_control.dashboard.services import DashboardData
from mission_control.planner.planner_generator import PlannerGenerator
from mission_control.workbook.base_sheet import BaseSheet
from mission_control.workbook.styles import solid_fill, thin_border


class DashboardSheet(BaseSheet):
    """Build the PGPEM Mission Control dashboard sheet."""

    title = "Dashboard"

    def __init__(
        self,
        *,
        planner_generator: PlannerGenerator | None = None,
        **kwargs: Any,
    ) -> None:
        super().__init__(**kwargs)
        self.planner_generator = planner_generator

    def build(self) -> None:
        """Create the formatted dashboard worksheet."""
        worksheet = self.worksheet
        self.set_column_widths(
            {
                "A": 20,
                "B": 16,
                "C": 16,
                "D": 16,
                "E": 14,
                "F": 18,
                "G": 16,
                "H": 16,
                "I": 16,
                "J": 16,
            }
        )
        self.prepare_sheet(
            title_text=self.app_config.project_name,
            subtitle=(
                f"{self.app_config.exam.name} command dashboard | "
                f"Generated {date.today():%d %b %Y}"
            ),
            merge_to="J1",
        )

        metrics = self._dashboard_data()
        self._write_kpis(metrics)
        self._write_progress_table()
        self._write_focus_table()
        self._write_chart()
        self.auto_size_columns()
        self._apply_page_setup()

    def _planner_generator(self) -> PlannerGenerator:
        if self.planner_generator is None:
            self.planner_generator = PlannerGenerator.from_app_config(
                app_config=self.app_config,
            )
        return self.planner_generator

    def _dashboard_data(self) -> DashboardData:
        if self.context is not None:
            return self.context.dashboard_data

        plan = self._planner_generator().generate()
        total_study_days = len(plan)
        total_study_hours = sum(day.study_hours for day in plan)
        total_revision_days = sum(1 for day in plan if day.revision)
        total_mock_tests = sum(1 for day in plan if day.mock_test)
        completed_days = sum(1 for day in plan if day.status == "Completed")
        completion_percentage = round((completed_days / total_study_days) * 100)

        return DashboardData(
            total_study_days=total_study_days,
            total_study_hours=total_study_hours,
            total_revision_days=total_revision_days,
            total_mock_tests=total_mock_tests,
            completion_percentage=completion_percentage,
        )

    def _write_kpis(self, metrics: DashboardData) -> None:
        metrics_dataframe = pd.DataFrame(
            [
                ("Total Study Days", metrics.total_study_days, "days"),
                ("Total Study Hours", metrics.total_study_hours, "hrs"),
                ("Revision Days", metrics.total_revision_days, "days"),
                ("Mock Tests", metrics.total_mock_tests, "tests"),
                ("Completion", metrics.completion_percentage, "%"),
            ],
            columns=["Metric", "Value", "Unit"],
        )

        start_columns = [1, 3, 5, 7, 9]
        for record, start_column in zip(
            metrics_dataframe.itertuples(index=False),
            start_columns,
        ):
            self._write_kpi_card(
                label=str(record.Metric),
                value=record.Value,
                unit=str(record.Unit),
                start_column=start_column,
            )

    def _write_kpi_card(
        self,
        *,
        label: str,
        value: int,
        unit: str,
        start_column: int,
    ) -> None:
        worksheet = self.worksheet
        worksheet.merge_cells(
            start_row=4,
            start_column=start_column,
            end_row=4,
            end_column=start_column + 1,
        )
        worksheet.merge_cells(
            start_row=5,
            start_column=start_column,
            end_row=6,
            end_column=start_column + 1,
        )

        label_cell = worksheet.cell(row=4, column=start_column, value=label)
        label_cell.fill = solid_fill(self.theme.neutral)
        label_cell.font = Font(
            name=self.theme.body_font,
            size=self.theme.body_size,
            bold=True,
            color=self.theme.muted_text,
        )
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        value_cell = worksheet.cell(
            row=5,
            column=start_column,
            value=f"{value} {unit}",
        )
        value_cell.fill = solid_fill(self.theme.white)
        value_cell.font = Font(
            name=self.theme.header_font,
            size=16,
            bold=True,
            color=self.theme.primary_dark,
        )
        value_cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in range(4, 7):
            for column in range(start_column, start_column + 2):
                cell = worksheet.cell(row=row, column=column)
                cell.border = thin_border(self.theme.border)

    def _write_progress_table(self) -> None:
        progress = pd.DataFrame(
            [
                ("Quantitative Aptitude", 42, 30, 71, "Practice arithmetic speed"),
                ("DILR", 30, 18, 60, "Review set selection patterns"),
                ("VARC", 36, 27, 75, "Increase RC review cadence"),
                ("Vocabulary", 20, 16, 80, "Maintain daily revision"),
                ("Mock Analysis", 12, 8, 67, "Log root causes after each mock"),
            ],
            columns=["Track", "Planned", "Completed", "Progress %", "Next Action"],
        )
        self.write_dataframe(progress, start_row=8)

        worksheet = self.worksheet
        for row in range(9, 14):
            progress_cell = worksheet.cell(row=row, column=4)
            progress_cell.number_format = '0"%"'
            progress_cell.alignment = Alignment(horizontal="center")

            if int(progress_cell.value) >= 75:
                progress_cell.fill = solid_fill(self.theme.secondary)
            elif int(progress_cell.value) >= 65:
                progress_cell.fill = solid_fill("FFF2CC")
            else:
                progress_cell.fill = solid_fill("FCE4D6")

    def _write_focus_table(self) -> None:
        focus = pd.DataFrame(
            [
                ("Mon-Wed", "Quant drills", "90 min/day", "Arithmetic + algebra"),
                ("Thu", "DILR sets", "2 sets", "Timed selection practice"),
                ("Fri", "VARC", "3 RC passages", "Review misses immediately"),
                ("Weekend", "Mock + analysis", "1 full mock", "Tag weak concepts"),
            ],
            columns=["Window", "Focus", "Target", "Review Note"],
        )

        label_cell = self.worksheet.cell(row=16, column=1, value="This Week's Focus")
        label_cell.font = Font(
            name=self.theme.header_font,
            size=self.theme.header_size,
            bold=True,
            color=self.theme.primary_dark,
        )
        self.write_dataframe(focus, start_row=17)

    def _write_chart(self) -> None:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Completion by Track"
        chart.y_axis.title = "Track"
        chart.x_axis.title = "Progress %"
        chart.height = 7
        chart.width = 10

        data = Reference(self.worksheet, min_col=4, min_row=8, max_row=13)
        categories = Reference(self.worksheet, min_col=1, min_row=9, max_row=13)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        self.worksheet.add_chart(chart, "G8")

    def _apply_page_setup(self) -> None:
        worksheet = self.worksheet
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
