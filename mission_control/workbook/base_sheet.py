"""Base abstractions for workbook sheets."""

from __future__ import annotations

from abc import ABC, abstractmethod
from collections.abc import Mapping
from typing import TYPE_CHECKING, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from mission_control.core.config import AppConfig, load_app_config
from mission_control.core.logger import get_logger
from mission_control.core.theme import Theme
from mission_control.workbook.styles import (
    apply_column_widths,
    apply_table_header_style,
    auto_size_columns,
    solid_fill,
    thin_border,
)

if TYPE_CHECKING:
    from mission_control.application.context import ApplicationContext


class BaseSheet(ABC):
    """Base class for generated workbook sheets."""

    title: str

    def __init__(
        self,
        *,
        app_config: AppConfig | None = None,
        theme: Theme | None = None,
    ) -> None:
        self.app_config = app_config or load_app_config()
        self.theme = theme or Theme()
        self.logger = get_logger(f"{__name__}.{self.__class__.__name__}")
        self.workbook: Workbook | None = None
        self._worksheet: Worksheet | None = None
        self.context: ApplicationContext | None = None

    @property
    def worksheet(self) -> Worksheet:
        """Return the attached worksheet."""
        if self._worksheet is None:
            raise RuntimeError(f"Sheet {self.title!r} is not attached to a workbook.")
        return self._worksheet

    def attach(self, workbook: Workbook) -> None:
        """Attach the sheet object to an openpyxl workbook."""
        self.workbook = workbook
        self._worksheet = self._create_or_replace_worksheet(self.title)

    @abstractmethod
    def build(self) -> None:
        """Build the sheet content and formatting."""

    def set_tab_color(self, color: str) -> None:
        """Set the worksheet tab color."""
        self.worksheet.sheet_properties.tabColor = color

    def set_title(
        self,
        text: str,
        *,
        cell: str = "A1",
        merge_to: str = "H1",
        subtitle: str | None = None,
    ) -> None:
        """Render a page title and optional subtitle."""
        self.worksheet.merge_cells(f"{cell}:{merge_to}")
        title_cell = self.worksheet[cell]
        title_cell.value = text
        title_cell.font = Font(
            name=self.theme.header_font,
            size=self.theme.title_size,
            bold=True,
            color=self.theme.primary_dark,
        )
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        if subtitle:
            subtitle_row = title_cell.row + 1
            subtitle_cell = self.worksheet.cell(row=subtitle_row, column=1)
            subtitle_cell.value = subtitle
            subtitle_cell.font = Font(
                name=self.theme.body_font,
                size=self.theme.subtitle_size,
                color=self.theme.muted_text,
            )

    def prepare_sheet(
        self,
        *,
        title_text: str | None = None,
        subtitle: str | None = None,
        merge_to: str = "H1",
    ) -> None:
        """Apply shared sheet chrome before writing content."""
        self.worksheet.sheet_view.showGridLines = False
        self.set_tab_color(self.theme.primary)
        self.set_title(
            title_text or self.title,
            merge_to=merge_to,
            subtitle=subtitle,
        )
        self.freeze_first_row()

    def freeze_first_row(self) -> None:
        """Freeze the first worksheet row."""
        self.freeze_at("A2")

    def freeze_header_row(self, header_row: int) -> None:
        """Freeze panes below the header row."""
        self.freeze_at(f"A{header_row + 1}")

    def write_table(
        self,
        headers: list[str],
        rows: list[tuple[Any, ...]],
        *,
        start_row: int = 4,
        start_column: int = 1,
    ) -> tuple[int, int]:
        """Write a table from headers and row tuples using shared styling."""
        dataframe = pd.DataFrame(rows, columns=headers)
        return self.write_dataframe(
            dataframe,
            start_row=start_row,
            start_column=start_column,
        )

    def write_dataframe(
        self,
        dataframe: pd.DataFrame,
        *,
        start_row: int,
        start_column: int = 1,
        header: bool = True,
    ) -> tuple[int, int]:
        """Write a pandas DataFrame and return its bottom-right cell."""
        row_offset = 0

        if header:
            for column_offset, column_name in enumerate(dataframe.columns):
                cell = self.worksheet.cell(
                    row=start_row,
                    column=start_column + column_offset,
                    value=str(column_name),
                )
                cell.border = thin_border(self.theme.border)
            apply_table_header_style(
                self.worksheet,
                start_row,
                start_column,
                start_column + len(dataframe.columns) - 1,
                self.theme,
            )
            row_offset = 1

        for dataframe_row in dataframe.itertuples(index=False):
            for column_offset, value in enumerate(dataframe_row):
                cell = self.worksheet.cell(
                    row=start_row + row_offset,
                    column=start_column + column_offset,
                    value=self._normalize_value(value),
                )
                cell.font = Font(
                    name=self.theme.body_font,
                    size=self.theme.body_size,
                    color=self.theme.text,
                )
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border(self.theme.border)
            row_offset += 1

        return (
            start_row + row_offset - 1,
            start_column + len(dataframe.columns) - 1,
        )

    def set_column_widths(self, widths: Mapping[str, float]) -> None:
        """Apply explicit column widths."""
        apply_column_widths(self.worksheet, widths)

    def auto_size_columns(self) -> None:
        """Auto-size worksheet columns based on content."""
        auto_size_columns(self.worksheet)

    def freeze_at(self, cell: str) -> None:
        """Freeze panes at the provided cell coordinate."""
        self.worksheet.freeze_panes = cell

    def apply_auto_filter(
        self,
        *,
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
    ) -> None:
        """Apply an Excel auto-filter to a rectangular range."""
        top_left = self.worksheet.cell(start_row, start_column).coordinate
        bottom_right = self.worksheet.cell(end_row, end_column).coordinate
        self.worksheet.auto_filter.ref = f"{top_left}:{bottom_right}"

    def apply_alternating_rows(
        self,
        *,
        start_row: int,
        end_row: int,
        start_column: int,
        end_column: int,
        fill_color: str | None = None,
    ) -> None:
        """Apply alternating row fills to a rectangular data range."""
        fill = solid_fill(fill_color or self.theme.neutral)
        for row in range(start_row, end_row + 1):
            if (row - start_row) % 2:
                continue
            self._apply_row_fill(row, start_column, end_column, fill)

    def _apply_row_fill(
        self,
        row: int,
        start_column: int,
        end_column: int,
        fill: PatternFill,
    ) -> None:
        for column in range(start_column, end_column + 1):
            self.worksheet.cell(row=row, column=column).fill = fill

    def _create_or_replace_worksheet(self, title: str) -> Worksheet:
        if self.workbook is None:
            raise RuntimeError("Cannot create a worksheet before attaching a workbook.")

        if title in self.workbook.sheetnames:
            existing = self.workbook[title]
            self.workbook.remove(existing)
        return self.workbook.create_sheet(title)

    @staticmethod
    def _normalize_value(value: Any) -> Any:
        if pd.isna(value):
            return None
        return value
