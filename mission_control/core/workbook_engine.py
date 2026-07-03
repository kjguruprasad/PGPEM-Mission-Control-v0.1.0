"""Workbook orchestration for PGPEM Mission Control."""

from __future__ import annotations

from pathlib import Path
from typing import TypeVar

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mission_control.core.config import WorkbookConfig
from mission_control.core.exceptions import SheetBuildError
from .logger import get_logger
from mission_control.workbook.base_sheet import BaseSheet


SheetT = TypeVar("SheetT", bound=BaseSheet)


class WorkbookEngine:
    """Create, build, and save the Mission Control Excel workbook."""

    def __init__(self, config: WorkbookConfig | None = None) -> None:
        self.config = config or WorkbookConfig()
        self.logger = get_logger(__name__)
        self.workbook = Workbook()
        self._configure_properties()
        self.logger.info("Workbook initialized")

    def create_sheet(self, name: str) -> Worksheet:
        """Create and return a raw worksheet."""
        return self.workbook.create_sheet(name)

    def remove_default_sheet(self) -> None:
        """Remove openpyxl's default sheet when it is still present."""
        sheet = self.workbook.active
        if sheet and sheet.title == "Sheet":
            self.workbook.remove(sheet)

    def add_sheet(self, sheet_type: type[SheetT]) -> SheetT:
        """Instantiate and build a concrete sheet class."""
        self.remove_default_sheet()
        sheet = sheet_type(self.workbook)

        try:
            sheet.build()
        except Exception as exc:  # pragma: no cover - defensive wrapper
            message = f"Failed to build sheet {sheet.title!r}"
            self.logger.exception(message)
            raise SheetBuildError(message) from exc

        self.logger.info("Sheet built: %s", sheet.title)
        return sheet

    def save(self) -> Path:
        """Persist the workbook and return the generated path."""
        self.config.output_dir.mkdir(parents=True, exist_ok=True)
        path = self.config.output_path
        self.workbook.save(path)
        self.logger.info("Workbook saved to %s", path)
        return path

    def _configure_properties(self) -> None:
        self.workbook.properties.creator = self.config.creator
        self.workbook.properties.title = self.config.title
        self.workbook.properties.subject = self.config.subject
