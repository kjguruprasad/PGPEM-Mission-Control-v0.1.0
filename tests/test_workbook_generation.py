from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from mission_control.core.config import WorkbookConfig
from mission_control.main import generate_workbook
from mission_control.planner.planner_generator import PlannerGenerator
from mission_control.workbook.base_sheet import BaseSheet
from mission_control.workbook.dashboard import DashboardSheet
from mission_control.workbook.sheets import (
    AnalyticsSheet,
    DailyPlannerSheet,
    DILRTrackerSheet,
    HabitsSheet,
    InterviewPrepSheet,
    MockTestSheet,
    NotesSheet,
    Planner106Sheet,
    QuantTrackerSheet,
    RevisionSheet,
    SettingsSheet,
    SmartGoalsSheet,
    VARCTrackerSheet,
    VocabularySheet,
)


EXPECTED_SHEETS = [
    "Dashboard",
    "SMART Goals",
    "106-Day Planner",
    "Revision Planner",
    "Daily Planner",
    "Quant Tracker",
    "DILR Tracker",
    "VARC Tracker",
    "Vocabulary",
    "Revision Tracker",
    "Mock Tests",
    "Analytics",
    "Habits",
    "Notes",
    "Interview Prep",
    "Settings",
]


def test_generate_workbook_creates_all_registered_sheets(tmp_path):
    config = WorkbookConfig(output_dir=tmp_path)

    workbook_path = generate_workbook(config)

    assert workbook_path == tmp_path / "PGPEM_Mission_Control.xlsx"
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == EXPECTED_SHEETS

    dashboard = workbook["Dashboard"]
    assert dashboard["A1"].value == "PGPEM Mission Control"
    assert dashboard["A4"].value == "Total Study Days"
    assert dashboard["A5"].value == "106 days"
    assert dashboard["A7"].value == "Total Revision Tasks"
    assert dashboard["A11"].value == "Track"
    assert dashboard["A12"].value == "Quantitative Aptitude"
    assert len(dashboard._charts) == 1


def test_dashboard_has_expected_formatting(tmp_path):
    workbook_path = generate_workbook(WorkbookConfig(output_dir=tmp_path))
    dashboard = load_workbook(workbook_path)["Dashboard"]

    assert dashboard.sheet_view.showGridLines is False
    assert dashboard.freeze_panes == "A2"
    assert dashboard["A11"].fill.fgColor.rgb == "001F4E78"
    assert dashboard.column_dimensions["A"].width >= 20


def test_standard_sheets_have_titles_headers_and_frozen_first_row(tmp_path):
    workbook_path = generate_workbook(WorkbookConfig(output_dir=tmp_path))
    workbook = load_workbook(workbook_path)

    for sheet_name in EXPECTED_SHEETS[1:]:
        worksheet = workbook[sheet_name]
        assert worksheet["A1"].value == sheet_name
        assert worksheet["A4"].value is not None
        expected_freeze = (
            "A5"
            if sheet_name in {"106-Day Planner", "Revision Planner"}
            else "A2"
        )
        assert worksheet.freeze_panes == expected_freeze
        assert worksheet.sheet_view.showGridLines is False


def test_yaml_config_drives_planner_and_settings(tmp_path):
    workbook_path = generate_workbook(WorkbookConfig(output_dir=tmp_path))
    workbook = load_workbook(workbook_path)

    planner = workbook["106-Day Planner"]
    settings = workbook["Settings"]

    assert planner["A110"].value == 106
    assert settings["B8"].value == 106


def test_planner_generator_creates_106_rows():
    plan = PlannerGenerator.from_files().generate()

    assert len(plan) == 106


def test_planner_generator_uses_correct_start_and_end_dates():
    plan = PlannerGenerator.from_files().generate()

    assert plan[0].date == date(2026, 7, 3)
    assert plan[-1].date == date(2026, 10, 16)


def test_planner_generator_applies_revision_frequency():
    plan = PlannerGenerator.from_files().generate()

    assert plan[6].revision == "Weekly Revision"
    assert plan[13].revision == "Weekly Revision"
    assert all(
        day.revision in {"Final Revision", "Light Revision"}
        for day in plan[-7:]
    )
    assert all(day.quant_topic == "" for day in plan[-7:])


def test_planner_generator_applies_mock_frequency():
    plan = PlannerGenerator.from_files().generate()
    mock_days = [day.day for day in plan if day.mock_test]

    assert mock_days == [14, 28, 42, 56, 70, 84, 98]


def test_all_sheet_classes_inherit_from_base_sheet():
    sheet_classes = [
        DashboardSheet,
        SmartGoalsSheet,
        Planner106Sheet,
        DailyPlannerSheet,
        QuantTrackerSheet,
        DILRTrackerSheet,
        VARCTrackerSheet,
        VocabularySheet,
        RevisionSheet,
        MockTestSheet,
        AnalyticsSheet,
        HabitsSheet,
        NotesSheet,
        InterviewPrepSheet,
        SettingsSheet,
    ]

    assert all(issubclass(sheet_class, BaseSheet) for sheet_class in sheet_classes)
